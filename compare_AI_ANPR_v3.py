# /!/usr/bin/python
# -*-coding:utf-8-*-

import sys
import os
import pandas as pd
import numpy as np
import argparse
import configparser
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Color
from difflib import SequenceMatcher
import datetime
from operator import itemgetter
import re
import math
from itertools import chain
from collections import defaultdict
# import distance
# import numpy.random.common
# import numpy.random.bounded_integers
# import numpy.random.entropy


def similarity_check(first_list, second_list, threshold=80):

    perfect_match = 0
    similar_match = 0
    similarity_info_list = []
    similar_idx_list = []

    if len(first_list) == len(second_list):
        for i in range(len(first_list)):
            sm_similarity = SequenceMatcher(None, first_list[i], second_list[i])
            sm_similarity_ratio = sm_similarity.ratio()*100
            # sor_similarity = (1 - distance.sorensen(check_ai, check_human))*100
            # jac_similarity = (1 - distance.jaccard(check_ai, check_human))*100
            # f1_similarity = f1_score(check_ai, check_human)
            similarity_info = [first_list[i], second_list[i], sm_similarity_ratio]
            similarity_info_list.append(similarity_info)
            if sm_similarity_ratio == 100:
                perfect_match += 1
                similar_idx_list.append(i)
            elif sm_similarity_ratio >= threshold:
                similar_match += 1
                similar_idx_list.append(i)

    return perfect_match, similar_match, similarity_info_list, similar_idx_list


def read_excel_file_info(anpr_excel_file, ai_excel_file, data_cfg):
    anpr_data = pd.read_excel(anpr_excel_file)
    for i in range(len(anpr_data.columns)):
        anpr_data.columns.values[i] = anpr_data.columns.values[i].replace(' ', '')
    ai_data = pd.read_excel(ai_excel_file)
    for i in range(len(ai_data.columns)):
        ai_data.columns.values[i] = ai_data.columns.values[i].replace(' ', '')

    columns = data_cfg['columns']
    camera_list = data_cfg['camera_id_list']['camera_id'].split(',\n')

    camera_id = columns['camera_id'].replace(' ', '')
    time = columns['time'].replace(' ', '')
    car_num = columns['car_num'].replace(' ', '')
    img_url = columns['img_url'].replace(' ', '')
    plate_url = columns['plate_url'].replace(' ', '')
    confidence = columns['confidence'].replace(' ', '')

    anpr_camera_id = anpr_data[camera_id]
    anpr_time = anpr_data[time]
    anpr_num = anpr_data[car_num]
    anpr_img_url = anpr_data[img_url]
    anpr_plate_url = anpr_data[plate_url]
    anpr_confidence = anpr_data[confidence]

    ai_camera_id = ai_data[camera_id]
    ai_time = ai_data[time]
    ai_num = ai_data[car_num]
    ai_img_url = ai_data[img_url]
    ai_plate_url = ai_data[plate_url]
    ai_confidence = ai_data[confidence]

    anpr_info_list_all = []
    ai_info_list_all = []

    for i in range(len(anpr_camera_id)):
        anpr_info_list = [anpr_camera_id[i], anpr_time[i], anpr_num[i], anpr_img_url[i], anpr_plate_url[i], anpr_confidence[i]]
        anpr_info_list_all.append(anpr_info_list)

    for i in range(len(ai_camera_id)):
        ai_info_list = [ai_camera_id[i], ai_time[i], ai_num[i], ai_img_url[i], ai_plate_url[i], ai_confidence[i]]
        ai_info_list_all.append(ai_info_list)

    return camera_list, anpr_info_list_all, ai_info_list_all


def find_match(anpr_num_list, ai_num_list):

    match_cnt = 0
    general_plate_num = 0
    general_plate_match_cnt = 0
    sales_plate_num = 0
    sales_plate_match_cnt = 0
    anpr_error_cnt = 0

    match_idx_list = []
    unmatch_idx_list = []

    #regular expression for inspecting anpr_num
    plt_gen = re.compile('[0-9]{2}[가-힣][0-9]{4}')
    plt_new = re.compile('[0-9]{3}[가-힣][0-9]{4}')
    plt_reg = re.compile('[가-힣]{2}[0-9]{2}[가-힣][0-9]{4}')
    plt_dip = re.compile('[가-힣]{2}[0-9]{3}[-][0-9]{3}')

    if len(ai_num_list) >= 1:
        for anpr_num in anpr_num_list:
            if not plt_gen.match(anpr_num) and not plt_new.match(anpr_num) and not plt_reg.match(anpr_num) and not plt_dip.match(anpr_num):
                anpr_error_cnt += 1
            else:
                try:
                    int(anpr_num[:2])  #일반 번호판
                    general_plate_num += 1
                    for ai_num in ai_num_list:
                        if anpr_num == ai_num:
                            match_idx_list.append(ai_num_list.index(ai_num))
                            match_cnt += 1
                            general_plate_match_cnt += 1
                            break
                except ValueError: #영업용
                    sales_plate_num += 1
                    for ai_num in ai_num_list:
                        if anpr_num == ai_num:
                            match_idx_list.append(ai_num_list.index(ai_num))
                            match_cnt += 1
                            sales_plate_match_cnt += 1
                            break

        for i in range(len(ai_num_list)):
            if i not in match_idx_list:
                unmatch_idx_list.append(i)

        return match_cnt, general_plate_num, general_plate_match_cnt, sales_plate_num, sales_plate_match_cnt, \
               anpr_error_cnt, match_idx_list, unmatch_idx_list

    else:
        for anpr_num in anpr_num_list:
            if not plt_gen.match(anpr_num) and not plt_new.match(anpr_num) and not plt_reg.match(anpr_num) and not plt_dip.match(anpr_num):
                anpr_error_cnt += 1
            else:
                try:
                    int(anpr_num[:2])  #일반 번호판
                    general_plate_num += 1
                except ValueError: #영업용, 외교
                    sales_plate_num += 1
        return general_plate_num, sales_plate_num, anpr_error_cnt


def compare_anpr_ai(anpr_info, ai_info):

    msg = ''
    if len(anpr_info) >= 1:
        if len(ai_info) >= 1:
            unmatch_list = []
            match_list = []
            anpr_detect_num = len(anpr_info)
            ai_detect_num = len(ai_info)
            msg += 'ANPR detected count : {}\n'.format(anpr_detect_num)
            msg += 'AI detected count : {}\n'.format(ai_detect_num)
            msg += 'Detect ratio : {:4.2f}%\n'.format((ai_detect_num/anpr_detect_num)*100)

            anpr_car_num_list = np.array(anpr_info).T.tolist()[2]
            ai_car_num_list = np.array(ai_info).T.tolist()[2]

            match_cnt, general_plate_num, general_plate_match_cnt, sales_plate_num, sales_plate_match_cnt\
                , anpr_error_cnt, match_idx_list, unmatch_idx_list = find_match(anpr_car_num_list, ai_car_num_list)

            msg += 'General plate count : {}\n'.format(general_plate_num)
            msg += 'General plate Match count : {}\n'.format(general_plate_match_cnt)
            general_plate_match_ratio = 0
            if general_plate_num > 0:
                general_plate_match_ratio = (general_plate_match_cnt/general_plate_num)*100
            msg += 'General plate Match ratio : {:4.2f}%\n'.format(general_plate_match_ratio)

            msg += 'Sales plate count : {}\n'.format(sales_plate_num)
            msg += 'Sales plate Match count : {}\n'.format(sales_plate_match_cnt)
            sales_plate_match_ratio = 0
            if sales_plate_num > 0:
                sales_plate_match_ratio = (sales_plate_match_cnt/sales_plate_num)*100
            msg += 'Sales plate Match ratio : {:4.2f}%\n'.format(sales_plate_match_ratio)

            msg += 'Match Count : {}\n'.format(len(match_idx_list))
            msg += 'Match ratio : {:4.2f}%\n'.format((match_cnt / (anpr_detect_num-anpr_error_cnt)) * 100)
            msg += 'ANPR error count : {}\n'.format(anpr_error_cnt)
            msg += '\n=========================================\n'

            #unmatch list 추가
            for unmatch_idx in unmatch_idx_list:
                unmatch_list.append(ai_info[unmatch_idx])

            for match_idx in match_idx_list:
                match_list.append(ai_info[match_idx])

            return msg, unmatch_list, match_list

        else:
            anpr_detect_num = len(anpr_info)
            anpr_car_num_list = np.array(anpr_info).T.tolist()[2]
            ai_car_num_list = ai_info
            general_plate_num, sales_plate_num, anpr_error_cnt = find_match(anpr_car_num_list, ai_car_num_list)
            msg += 'ANPR detected count : {}\n'.format(anpr_detect_num)
            msg += 'AI detected count : None\n'
            msg += 'Detect ratio : None\n'
            msg += 'General plate count : {}\n'.format(general_plate_num)
            msg += 'General plate Match count : None\n'
            msg += 'General plate Match ratio : None\n'
            msg += 'Sales plate count : {}\n'.format(sales_plate_num)
            msg += 'Sales plate Match count : None\n'
            msg += 'Sales plate Match ratio : None\n'
            msg += 'Match Count : None\n'
            msg += 'Match ratio : None\n'
            msg += 'ANPR error count : {}\n'.format(anpr_error_cnt)
            msg += '\n=========================================\n'

            return msg

    else:
        if len(ai_info) >= 1:
            ai_detect_num = len(ai_info)
            msg += 'ANPR detected count : None\n'
            msg += 'AI detected count : {}\n'.format(ai_detect_num)
            msg += 'Detect ratio : None\n'
            msg += 'General plate count : None\n'
            msg += 'General plate Match count : None\n'
            msg += 'General plate Match ratio : None\n'
            msg += 'Sales plate count : None\n'
            msg += 'Sales plate Match count : None\n'
            msg += 'Sales plate Match ratio : None\n'
            msg += 'Match Count : None\n'
            msg += 'Match ratio : None\n'
            msg += 'ANPR error count : None\n'
            msg += '\n=========================================\n'

            return msg

        else:
            msg += 'ANPR detected count : None\n'
            msg += 'AI detected count : None\n'
            msg += 'Detect ratio : None\n'
            msg += 'General plate count : None\n'
            msg += 'General plate Match count : None\n'
            msg += 'General plate Match ratio : None\n'
            msg += 'Sales plate count : None\n'
            msg += 'Sales plate Match count : None\n'
            msg += 'Sales plate Match ratio : None\n'
            msg += 'Match Count : None\n'
            msg += 'Match ratio : None\n'
            msg += 'ANPR error count : None\n'
            msg += '\n=========================================\n'

            return msg


def stats_per_confidence_level(info_list, conf_col=-1, conf_interval=10, delete_level=False):
    level_num = math.ceil(100/conf_interval)
    level_dict = {}
    for i in range(level_num):
        if i == 0:
            level_dict[i] = [conf_interval*i, conf_interval*(i+1)]
            level_dict[i] = [conf_interval*i, conf_interval*(i+1)]
        else:
            level_dict[i] = [conf_interval*i+1, conf_interval*(i+1)]
            level_dict[i] = [conf_interval*i+1, conf_interval * (i + 1)]

    for key, value in level_dict.items():
        match_cnt = 0
        c_min = value[0]
        c_max = value[1]
        for info in info_list:
            conf_level = info[conf_col]
            if c_min <= conf_level <= c_max:
                match_cnt += 1
        level_dict[key].append(match_cnt)

    if delete_level:
        for key, value in level_dict.items():
            del value[0]
            del value[0]

    return level_dict


def write_excel_file(title, list_for_excel=None, dict_for_excel=None, msg=None):

    if list_for_excel:

        wb = Workbook()
        ws = wb.active
        for contents in list_for_excel:
            if isinstance(contents, list):
                ws.append(contents)
                wb.save(title)
            else:
                ws.append(list_for_excel)
                wb.save(title)

    if dict_for_excel:

        if os.path.isfile(title):
            wb = load_workbook(title)
            ws = wb.active

        else:
            wb = Workbook()
            ws = wb.active
            header_list = ['신뢰도 min', '신뢰도 max','개수','정답 수','인식률']
            for i in range(len(header_list)):
                ws.cell(row=1, column=i+1).fill = PatternFill(patternType='solid', fgColor=Color('FFC000'))
                ws.cell(row=1, column=i+1).value = header_list[i]

        for key, value in dict_for_excel.items():
            ws.append(value)

        wb.save(title)

    if msg:
        write_list = []
        msg_list = msg.split('\n')
        for x in msg_list:
            if x.find(":") == -1:
                continue
            else:
                x = x.split(':')[1]
                write_list.append(x)

        if os.path.isfile(title):
            wb = load_workbook(title)
            ws = wb.active
            ws.append(write_list)
            wb.save(title)

        else:
            wb = Workbook()
            ws = wb.active
            header_list = ['측정일', 'ID', '시작시간', '종료시간', 'ANPR 검지차량', 'AI 검지차량', 'ANPR대비 검지율',
                           '일반 번호판 개수', '일반 번호판 일치 수', '일치율', '지역 번호판 개수', '지역 번호판 일치 수',
                           '일치율', '번호판 일치 수(전체)', 'ANPR 기준 일치율(전체)', 'ANPR 오류 개수']
            for i in range(len(header_list)):
                ws.cell(row=1, column=i+1).fill = PatternFill(patternType='solid', fgColor=Color('FFC000'))
                ws.cell(row=1, column=i+1).value = header_list[i]

            ws.append(write_list)
            wb.save(title)


def main():
    now = datetime.datetime.now()
    anpr_path = input("Type ANPR excel path: ")
    AI_path = input("Type AI excel path: ")
    Save_Unmatch_list = input("Save Unmatch list(Y/N) : ")

    parser = argparse.ArgumentParser()
    parser.add_argument("--ini_file", required=True, help="ini file for data information")
    args = parser.parse_args()
    data_cfg = configparser.ConfigParser()
    data_cfg.read(args.ini_file, encoding='utf-8')

    camera_list, anpr_info_list_all, ai_info_list_all = read_excel_file_info(anpr_path, AI_path, data_cfg)
    anpr_info_list_all = sorted(anpr_info_list_all, key=itemgetter(0))
    ai_info_list_all = sorted(ai_info_list_all, key=itemgetter(0))
    total_conf_dict = stats_per_confidence_level(ai_info_list_all)

    total_match_list = []

    for idx, cam_id in enumerate(camera_list):
        msg = ''
        anpr_info_list_sep = []
        ai_info_list_sep = []

        for anpr_info in anpr_info_list_all:
            if anpr_info[0] == cam_id:
                anpr_info_list_sep.append(anpr_info)

        for ai_info in ai_info_list_all:
            if ai_info[0] == cam_id:
                ai_info_list_sep.append(ai_info)

        if len(anpr_info_list_sep) >= 1:
            if len(ai_info_list_sep) >= 1:
                msg += '#{}\n'.format(idx)
                msg += 'Date : {}/{}\n'.format(now.month, now.day)
                msg += 'CAMERA ID : {}\n'.format(cam_id)
                msg += 'Start time : {}\n'.format(anpr_info_list_sep[0][1])
                msg += 'End time : {}\n'.format(anpr_info_list_sep[-1][1])
                compare_msg, unmatch_list, match_list = compare_anpr_ai(anpr_info_list_sep, ai_info_list_sep)
                total_match_list.extend(match_list)
                msg += compare_msg
                print(msg)
                excel_title = '{}_{}'.format(anpr_info_list_sep[0][1].split(' ')[0],
                                             anpr_info_list_sep[0][1].split(' ')[1].split(':')[0])
                write_excel_file(title='Detect info.xlsx', msg=msg)

                if Save_Unmatch_list == 'Y' or Save_Unmatch_list == 'y':
                    write_excel_file(title='Unmatch list({}_{}).xlsx'.format(cam_id, excel_title),
                                     list_for_excel=unmatch_list)


            else:
                msg += '#{}\n'.format(idx)
                msg += 'Date : {}/{}\n'.format(now.month, now.day)
                msg += 'CAMERA ID : {}\n'.format(cam_id)
                msg += 'Start time : {}\n'.format(anpr_info_list_sep[0][1])
                msg += 'End time : {}\n'.format(anpr_info_list_sep[-1][1])
                compare_msg = compare_anpr_ai(anpr_info_list_sep, ai_info_list_sep)
                msg += compare_msg
                print(msg)
                write_excel_file(title='Detect info.xlsx', msg=msg)
                if Save_Unmatch_list == 'Y' or Save_Unmatch_list == 'y':
                    print("Nothing to save, No information in ANPR excel file")

        else:
            if len(ai_info_list_sep) >= 1:
                msg += '#{}\n'.format(idx)
                msg += 'Date : {}/{}\n'.format(now.month, now.day)
                msg += 'CAMERA ID : {}\n'.format(cam_id)
                msg += 'Start time : {}\n'.format(ai_info_list_sep[0][1])
                msg+= 'End time : {}\n'.format(ai_info_list_sep[-1][1])
                compare_msg = compare_anpr_ai(anpr_info_list_sep, ai_info_list_sep)
                msg += compare_msg
                print(msg)
                write_excel_file(title='Detect info.xlsx', msg=msg)
                if Save_Unmatch_list == 'Y' or Save_Unmatch_list == 'y':
                    print("Nothing to save, No information in ANPR excel file")

            else:
                msg += '#{}\n'.format(idx)
                msg += 'Date : {}/{}\n'.format(now.month, now.day)
                msg += 'CAMERA ID : {}\n'.format(cam_id)
                msg += 'Start time : None\n'
                msg += 'End time : None\n'
                compare_msg = compare_anpr_ai(anpr_info_list_sep, ai_info_list_sep)
                msg += compare_msg
                print(msg)
                write_excel_file(title='Detect info.xlsx', msg=msg)
                if Save_Unmatch_list == 'Y' or Save_Unmatch_list == 'y':
                    print("Nothing to save, No information in ANPR excel file")

    match_conf_dict = stats_per_confidence_level(total_match_list, delete_level=True)

    final_conf_dict = defaultdict(list)
    for key, value in chain(total_conf_dict.items(), match_conf_dict.items()):
        final_conf_dict[key].extend(value)

    for key, value in final_conf_dict.items():
        total = value[-2]
        match = value[-1]
        rate = str(round((match/total)*100, 2)) + '%'
        final_conf_dict[key].append(rate)

    write_excel_file(title='신뢰도구간별인식률.xlsx', dict_for_excel=final_conf_dict)


if __name__ == "__main__":

    if len(sys.argv) == 1:

        sys.argv.extend([
                         "--ini_file", "data_info.ini",
                         ])

    main()

