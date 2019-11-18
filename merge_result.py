# /!/usr/bin/python
# -*-coding:utf-8-*-

import os
import sys
import pandas as pd
import argparse
import configparser
import math
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Color, Font, Alignment
import traceback


def read_excel_file_info(input_dir, data_cfg):

    columns = data_cfg['columns']
    camera_id = columns['camera_id']
    confidence = columns['confidence']
    answer = columns['answer']

    camera_id_list = []
    result_info_list = []
    file_list = os.listdir(input_dir)
    for file in file_list:
        if file.split('.')[-1] == 'xlsx':
            result_data = pd.read_excel(os.path.join(input_dir, file))
            result_cam_id = result_data[camera_id]
            result_conf = result_data[confidence]
            result_ans = result_data[answer]

            for i in range(result_data.shape[0]):
                if str(result_cam_id[i])[0] == 'A':
                    if result_cam_id[i] not in camera_id_list:
                        camera_id_list.append(result_cam_id[i])
                    result_info_list.append([result_cam_id[i], result_conf[i], result_ans[i]])

    return camera_id_list, result_info_list


def make_dict_by_cam_id(info_list, id_col_num=0, conf_col_num=1, ans_col_num=2):
    cam_id_dict = {}
    for info in info_list:
        cam_id = info[id_col_num]
        confidence = info[conf_col_num]
        answer = info[ans_col_num]
        if cam_id in cam_id_dict:
            cam_id_dict[cam_id].append([confidence, answer])
        else:
            cam_id_dict[cam_id] = [[confidence, answer]]

    return cam_id_dict


def make_dict_for_total(info_list, conf_col_num=1, ans_col_num=2):
    total_dict = {}
    for info in info_list:
        confidence = info[conf_col_num]
        answer = info[ans_col_num]
        if 'TOTAL' in total_dict:
            total_dict['TOTAL'].append([confidence, answer])
        else:
            total_dict['TOTAL'] = [[confidence, answer]]

    return total_dict


def make_dict_per_confidence_level(info_list, conf_col=0, conf_interval=10, interval_apply=80):
    level_num = math.ceil(interval_apply / conf_interval)
    level_dict = {}

    for i in range(level_num):
        if i == 0:
            level_dict[i] = [conf_interval * i, conf_interval * (i + 1)]
        else:
            level_dict[i] = [conf_interval * i + 1, conf_interval * (i + 1)]

    for i in range(interval_apply, 100):
        level_dict[i] = [i, i + 1]

    for key, val in level_dict.items():
        c_min = val[0]
        c_max = val[1]
        for info in info_list:
            conf_level = info[conf_col]
            try:
                if c_min <= conf_level <= c_max:
                    level_dict[key].extend(info[1])
            except Exception as e1:
                print("exception1:{}".format(e1))
                print(traceback.format_exc())

    return level_dict


def dict_stats(level_dict):
    level_stats_dict = {}
    for key, val in level_dict.items():
        c_min = val[0]
        c_max = val[1]
        O_cnt = 0
        X_cnt = 0
        q_cnt = 0
        XX_cnt = 0
        for ans in val:
            if ans == 'O':
                O_cnt += 1
            if ans == 'X':
                X_cnt += 1
            if ans == 'q':
                q_cnt += 1
            if ans == 'XX':
                XX_cnt += 1
        total_cnt = O_cnt + X_cnt + q_cnt + XX_cnt
        level_stats_dict[key] = [c_min, c_max, O_cnt, X_cnt, q_cnt, XX_cnt, total_cnt]

    return level_stats_dict


def make_excel_file(title, level_dict):
    if os.path.isfile(title):
        wb = load_workbook(title)
        ws = wb.active

    else:
        wb = Workbook()
        ws = wb.active
        header_list = ['카메라 ID', '신뢰도 min', '신뢰도 max', 'O(맞음)', 'X(틀림)', 'q(오검지)', 'XX(인식불가)', '합계']
        for i in range(len(header_list)):
            ws.cell(row=1, column=i + 1).fill = PatternFill(patternType='solid', fgColor=Color('FFC000'))
            ws.cell(row=1, column=i + 1).value = header_list[i]

    key_cnt = 0
    for key, val_dict in level_dict.items():
        key_start = key_cnt + 2
        for f in val_dict.values():
            list_for_excel = []
            list_for_excel.append(key)
            list_for_excel.extend(f)
            key_cnt += 1
            ws.append(list_for_excel)
        key_end = key_cnt + 1
        #merge camera ID_cells
        merge_start = 'A' + str(key_start)
        merge_end = 'A' + str(key_end)
        merge_range = ':'.join([merge_start, merge_end])
        ws.merge_cells(merge_range)
        ws[merge_start] = key
        ws[merge_start].font = Font(size=13, bold=True)
        ws[merge_start].alignment = Alignment(horizontal='center', vertical='center')
        #merge confidence cells
        for i in range(key_start, key_end+1):
            merge_start ='B'+str(i)
            merge_end = 'C'+str(i)
            start_val = ws.cell(i, 2).value
            end_val = ws.cell(i, 3).value
            merge_range = ':'.join([merge_start, merge_end])
            ws.merge_cells(merge_range)
            ws[merge_start] = '-'.join([str(start_val), str(end_val)])
            ws[merge_start].alignment = Alignment(horizontal='center', vertical='center')
    wb.save(title)


def main():
    input_dir = input("Input Directory : ")
    parser = argparse.ArgumentParser()
    parser.add_argument("--ini_file", required=True, help="ini file for data information")
    args = parser.parse_args()
    data_cfg = configparser.ConfigParser()
    data_cfg.read(args.ini_file, encoding='utf-8')

    final_dict = {}
    camera_id_list, result_info_list = read_excel_file_info(input_dir, data_cfg)
    cam_id_dict = make_dict_by_cam_id(result_info_list)
    total_dict = make_dict_for_total(result_info_list)
    for key, val in cam_id_dict.items():
        level_dict = make_dict_per_confidence_level(val)
        level_stats_dict = dict_stats(level_dict)
        final_dict[key] = level_stats_dict
    for key, val in total_dict.items():
        level_dict_total = make_dict_per_confidence_level(val)
        level_stats_dict_total = dict_stats(level_dict_total)
        final_dict[key] = level_stats_dict_total

    make_excel_file(title="신뢰도 구간별 통계.xlsx", level_dict=final_dict)


if __name__ == '__main__':

    if len(sys.argv) == 1:
        sys.argv.extend([
            "--ini_file", "result_excel_info.ini"
        ])
    main()
